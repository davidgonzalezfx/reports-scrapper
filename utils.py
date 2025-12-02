"""Utility functions for the reports scraper application.

This module provides file operations, report processing, and data aggregation
utilities for the educational platform scraper.
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

import pandas as pd
from openpyxl import Workbook

# Import new helper modules
from constants import (
    REPORTS_DIR, REPORT_TYPES, STUDENT_USAGE_COL_LISTEN,
    STUDENT_USAGE_COL_READ, STUDENT_USAGE_COL_QUIZ,
    STUDENT_USAGE_COL_STUDENT_NAME, STUDENT_USAGE_COL_CLASSROOM,
    STUDENT_USAGE_COL_INTERACTIVITY, STUDENT_USAGE_COL_PRACTICE_RECORDING,
    SKILL_COL_SKILL_NAME, SKILL_COL_CORRECT, SKILL_COL_TOTAL,
    SKILL_COL_ACCURACY, SKILL_COL_DATA_BAR, LEVEL_UP_COL_STUDENT,
    LEVEL_UP_COL_LEVEL, LEVEL_UP_COL_PROGRESS, SUMMARY_LABEL_TEACHERS,
    SUMMARY_LABEL_STUDENTS, SUMMARY_LABEL_LISTENS, SUMMARY_LABEL_READS,
    SUMMARY_LABEL_QUIZZES, SUMMARY_LABEL_TOTAL, TIMESTAMP_FORMAT,
    DANGEROUS_FILENAME_PATTERNS
)
from path_helpers import (
    get_base_path, is_frozen, get_executable_dir,
    get_all_reports_directories
)
from excel_helpers import (
    apply_header_style, auto_adjust_column_widths, add_data_bars,
    make_unique_sheet_name, create_summary_row, style_summary_row,
    create_separator_row, clean_row_data
)
from date_helpers import get_current_timestamp
from exceptions import FileOperationError, ReportProcessingError

logger = logging.getLogger(__name__)


def load_json(file_path: Union[str, Path], default: Any = None) -> Any:
    """Load JSON data from file with error handling.

    Args:
        file_path: Path to JSON file
        default: Default value to return if file doesn't exist or fails to load

    Returns:
        Loaded JSON data or default value
    """
    file_path = Path(file_path)

    if not file_path.exists():
        logger.info(f"JSON file not found: {file_path}")
        return default

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        logger.debug(f"Successfully loaded JSON from {file_path}")
        return data
    except (json.JSONDecodeError, OSError) as e:
        logger.error(f"Failed to load JSON from {file_path}: {e}")
        return default


def save_json(data: Any, file_path: Union[str, Path], indent: int = 2) -> bool:
    """Save data to JSON file with error handling.

    Args:
        data: Data to save
        file_path: Path to JSON file
        indent: JSON indentation level

    Returns:
        True if save successful, False otherwise
    """
    file_path = Path(file_path)

    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        logger.debug(f"Successfully saved JSON to {file_path}")
        return True
    except (OSError, TypeError) as e:
        logger.error(f"Failed to save JSON to {file_path}: {e}")
        return False


def convert_csv_to_xlsx(
    csv_path: Union[str, Path],
    remove_csv: bool = True
) -> Optional[str]:
    """Convert CSV file to XLSX format.

    Args:
        csv_path: Path to CSV file
        remove_csv: Whether to remove the original CSV file

    Returns:
        Path to XLSX file if successful, None otherwise
    """
    csv_path = Path(csv_path)
    xlsx_path = csv_path.with_suffix(".xlsx")

    try:
        logger.debug(f"Converting {csv_path} to XLSX")
        df = pd.read_csv(csv_path)
        df.to_excel(xlsx_path, index=False)

        logger.info(f"Successfully converted to XLSX: {xlsx_path.name}")

        if remove_csv:
            try:
                csv_path.unlink()
                logger.debug(f"Removed original CSV: {csv_path}")
            except OSError as e:
                logger.warning(f"Could not remove CSV {csv_path}: {e}")

        return str(xlsx_path)

    except Exception as e:
        logger.error(f"Failed to convert {csv_path} to XLSX: {e}")
        return None


def validate_filename(filename: str) -> bool:
    """Validate filename for security (prevent path traversal).

    Args:
        filename: Filename to validate

    Returns:
        True if filename is safe, False otherwise
    """
    if not filename:
        return False

    return not any(
        pattern in filename
        for pattern in DANGEROUS_FILENAME_PATTERNS
    )


def validate_user_data(user: Dict[str, Any]) -> bool:
    """Validate user data structure.

    Args:
        user: User dictionary to validate

    Returns:
        True if user data is valid, False otherwise
    """
    if not isinstance(user, dict):
        return False

    required_fields = ["username", "password"]
    return all(field in user and user[field] for field in required_fields)


def get_report_files(directory: Union[str, Path]) -> List[str]:
    """Get list of XLSX report files in directory.

    Args:
        directory: Path to reports directory

    Returns:
        List of XLSX filenames sorted by date (newest first)
    """
    import sys
    import os

    # Handle PyInstaller executable case
    if is_frozen():
        # When running as executable, check both internal and external
        # reports directories
        base_path = get_base_path()
        external_dir = Path(
            os.path.join(os.path.dirname(sys.executable), str(directory))
        )
        internal_dir = base_path / directory

        all_files = []

        # Check external directory first (priority for user downloads)
        if external_dir.exists():
            try:
                files = [
                    f.name for f in external_dir.glob("*.xlsx")
                    if f.is_file()
                ]
                all_files.extend(files)
                logger.debug(
                    f"Found {len(files)} report files in external "
                    f"directory: {external_dir}"
                )
            except OSError as e:
                logger.warning(
                    f"Error reading external reports directory "
                    f"{external_dir}: {e}"
                )

        # Check internal directory
        if internal_dir.exists():
            try:
                files = [
                    f.name for f in internal_dir.glob("*.xlsx")
                    if f.is_file()
                ]
                all_files.extend(files)
                logger.debug(
                    f"Found {len(files)} report files in internal "
                    f"directory: {internal_dir}"
                )
            except OSError as e:
                logger.warning(
                    f"Error reading internal reports directory "
                    f"{internal_dir}: {e}"
                )

        # Remove duplicates and sort
        unique_files = list(set(all_files))
        unique_files.sort(reverse=True)

        logger.debug(f"Total found {len(unique_files)} unique report files")
        return unique_files

    else:
        # Running in development mode - use original logic
        directory = Path(directory)

        if not directory.exists():
            logger.warning(f"Reports directory not found: {directory}")
            return []

        try:
            files = [f.name for f in directory.glob("*.xlsx") if f.is_file()]
            files.sort(reverse=True)  # Sort newest first

            logger.debug(f"Found {len(files)} report files")
            return files

        except OSError as e:
            logger.error(f"Error reading reports directory: {e}")
            return []


def _resolve_reports_directory(directory: Union[str, Path]) -> Optional[Path]:
    """Resolve reports directory path with PyInstaller support.

    This function handles both development and PyInstaller frozen executable modes.
    For frozen apps, it checks both external (next to .exe) and internal (_MEIPASS)
    directories, preferring the external directory where reports are downloaded.

    Args:
        directory: Path to reports directory (relative or absolute)

    Returns:
        First existing Path found, or None if no directory exists
    """
    import os

    # Handle absolute paths directly
    if os.path.isabs(str(directory)):
        abs_path = Path(directory)
        if abs_path.exists():
            logger.debug(f"Using absolute reports directory: {abs_path}")
            return abs_path
        logger.warning(f"Absolute reports directory not found: {abs_path}")
        return None

    # For relative paths, check all possible locations
    if is_frozen():
        # PyInstaller mode: check both external and internal directories
        # External directory (next to .exe) is checked first as it's where
        # reports are actually downloaded
        external_dir = get_executable_dir() / directory
        internal_dir = get_base_path() / directory

        # Prefer external directory (where reports are written)
        if external_dir.exists():
            logger.debug(f"Using external reports directory: {external_dir}")
            return external_dir

        # Fallback to internal directory
        if internal_dir.exists():
            logger.debug(f"Using internal reports directory: {internal_dir}")
            return internal_dir

        logger.warning(
            f"Reports directory not found in external ({external_dir}) "
            f"or internal ({internal_dir}) locations"
        )
        return None
    else:
        # Development mode: resolve against base path
        reports_directory = get_base_path() / directory

        if not reports_directory.exists():
            logger.warning(f"Reports directory not found: {reports_directory}")
            return None

        logger.debug(f"Using development reports directory: {reports_directory}")
        return reports_directory


def _create_combined_sheet(
    combined_wb: Workbook,
    report_type: str,
    files: List[Path]
) -> Optional[str]:
    """Create a combined sheet for a specific report type.

    Args:
        combined_wb: Workbook to add sheet to
        report_type: Type of report
        files: List of file paths for this report type

    Returns:
        Sheet name if successful, None otherwise
    """
    try:
        logger.debug(f"Processing {report_type} reports")

        # Create unique sheet name
        existing_names = [ws.title for ws in combined_wb.worksheets]
        sheet_name = make_unique_sheet_name(report_type, existing_names)
        ws = combined_wb.create_sheet(title=sheet_name)

        # Process all files and combine data
        all_data = []
        headers_written = False
        header_row_num = None
        separator_rows = []

        for file_path in sorted(files):
            try:
                # Validate file
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                # Read the Excel file
                try:
                    df = pd.read_excel(file_path, engine="openpyxl")
                except Exception as read_error:
                    logger.error(
                        f"Failed to read Excel file {file_path}: {read_error}"
                    )
                    continue

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Write headers only once
                if not headers_written:
                    ws.append(list(df.columns))
                    header_row_num = ws.max_row
                    headers_written = True

                # Extract username from filename
                filename = file_path.name
                username = filename.split("_")[0] if "_" in filename else filename

                # Create separator row
                num_cols = len(df.columns)
                separator_row = create_separator_row(username, num_cols)
                all_data.append(separator_row)
                separator_rows.append(len(all_data) - 1)

                # Write data rows
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))
                    all_data.append(cleaned_row)

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        # Write all collected data to sheet
        if all_data:
            _write_data_to_sheet(
                ws,
                all_data,
                header_row_num,
                separator_rows
            )

            # Add report-specific formatting
            _apply_report_specific_formatting(ws, report_type, all_data, files)

            # Auto-adjust column widths
            auto_adjust_column_widths(ws)

            logger.debug(f"Successfully added sheet for {report_type}")
            return sheet_name

        return None

    except Exception as e:
        logger.error(f"Unexpected error processing {report_type} reports: {e}")
        return None


def _write_data_to_sheet(
    ws,
    all_data: List[List],
    header_row_num: Optional[int],
    separator_rows: List[int]
) -> None:
    """Write data to worksheet with proper styling.

    Args:
        ws: Worksheet to write to
        all_data: List of data rows
        header_row_num: Header row number
        separator_rows: List of separator row indices
    """
    try:
        # Write all rows
        for row_data in all_data:
            ws.append(row_data)

        # Apply header styling
        if header_row_num:
            try:
                apply_header_style(ws, header_row_num)

                # Apply styling to separator rows
                for sep_row_idx in separator_rows:
                    sep_row_num = header_row_num + 1 + sep_row_idx
                    if sep_row_num <= ws.max_row:
                        apply_header_style(ws, sep_row_num)

            except Exception as style_error:
                logger.debug(f"Error applying header styling: {style_error}")

    except Exception as write_error:
        logger.error(f"Error writing data to sheet: {write_error}")
        raise


def _apply_report_specific_formatting(
    ws,
    report_type: str,
    all_data: List[List],
    files: List[Path]
) -> None:
    """Apply formatting specific to report type.

    Args:
        ws: Worksheet to format
        report_type: Type of report
        all_data: Data written to sheet
        files: Source files for the report
    """
    if report_type == "Student Usage":
        _add_student_usage_summary(ws, all_data, files)
    elif report_type == "Skill":
        _add_skill_data_bars(ws)


def _add_student_usage_summary(
    ws,
    all_data: List[List],
    files: List[Path]
) -> None:
    """Add summary rows to Student Usage sheet.

    Args:
        ws: Worksheet to add summary to
        all_data: Data in the sheet
        files: Source files
    """
    try:
        # Track separator rows for exclusion from calculations
        separator_rows = [
            i for i, row in enumerate(all_data)
            if row and str(row[0]).startswith("User:")
        ]

        # Calculate totals
        total_teachers = len(files)
        total_students = sum(
            1 for i in range(len(all_data))
            if i not in separator_rows
        )
        total_listens = 0
        total_reads = 0
        total_quizzes = 0

        # Process each data row (excluding separator rows)
        for i, row_data in enumerate(all_data):
            if i in separator_rows:
                continue

            # Sum listens
            if (len(row_data) > STUDENT_USAGE_COL_LISTEN and
                    pd.notna(row_data[STUDENT_USAGE_COL_LISTEN])):
                try:
                    total_listens += float(row_data[STUDENT_USAGE_COL_LISTEN])
                except (ValueError, TypeError):
                    pass

            # Sum reads
            if (len(row_data) > STUDENT_USAGE_COL_READ and
                    pd.notna(row_data[STUDENT_USAGE_COL_READ])):
                try:
                    total_reads += float(row_data[STUDENT_USAGE_COL_READ])
                except (ValueError, TypeError):
                    pass

            # Sum quizzes
            if (len(row_data) > STUDENT_USAGE_COL_QUIZ and
                    pd.notna(row_data[STUDENT_USAGE_COL_QUIZ])):
                try:
                    total_quizzes += float(row_data[STUDENT_USAGE_COL_QUIZ])
                except (ValueError, TypeError):
                    pass

        # Calculate total activities
        total_activities = total_listens + total_reads + total_quizzes

        # Get number of columns
        num_cols = ws.max_column if ws.max_column else 8

        # Add empty row for spacing
        ws.append([""] * num_cols)

        # Add summary rows
        summary_rows = [
            (SUMMARY_LABEL_TEACHERS, total_teachers),
            (SUMMARY_LABEL_STUDENTS, total_students),
            (SUMMARY_LABEL_LISTENS, int(total_listens)),
            (SUMMARY_LABEL_READS, int(total_reads)),
            (SUMMARY_LABEL_QUIZZES, int(total_quizzes)),
            (SUMMARY_LABEL_TOTAL, int(total_activities))
        ]

        for label, value in summary_rows:
            row = create_summary_row(label, value, num_cols)
            ws.append(row)
            style_summary_row(ws, ws.max_row)

        logger.debug(
            f"Added totals row to Student Usage sheet: "
            f"Students={total_students}, Teachers={total_teachers}, "
            f"Listens={total_listens}, Reads={total_reads}, "
            f"Quizzes={total_quizzes}"
        )

    except Exception as totals_error:
        logger.debug(
            f"Error adding totals row to Student Usage sheet: {totals_error}"
        )


def _add_skill_data_bars(ws) -> None:
    """Add data bars to Skill sheet.

    Args:
        ws: Worksheet to add data bars to
    """
    try:
        if ws.max_row > 1:  # Check if we have data rows beyond header
            # Add data bars to column D (Accuracy column)
            add_data_bars(
                ws,
                "D",
                start_row=2,
                end_row=ws.max_row,
                min_value=0,
                max_value=100
            )
            logger.debug(f"Added data bars to Skill sheet column D")

    except Exception as bar_error:
        logger.debug(f"Error adding data bars to Skill sheet: {bar_error}")


def combine_all_reports(directory: Union[str, Path]) -> Optional[str]:
    """Combine all reports by type into a single multi-sheet XLSX file.

    Args:
        directory: Path to the reports directory (can be relative or absolute)

    Returns:
        Path to the combined file if successful, None otherwise
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find all report files for each type
        reports_by_type = {}
        for report_type in REPORT_TYPES:
            # Use glob pattern to find files containing the report type
            pattern = f"*_{report_type}_*.xlsx"
            files = list(reports_directory.glob(pattern))
            if files:
                reports_by_type[report_type] = files
                logger.info(f"Found {len(files)} {report_type} reports")

        if not reports_by_type:
            logger.info("No reports found to combine")
            return None

        # Create a new workbook for combined reports
        combined_wb = Workbook()
        combined_wb.remove(combined_wb.active)  # Remove default sheet

        # Process each report type
        successful_sheets = 0
        for report_type, files in reports_by_type.items():
            sheet_name = _create_combined_sheet(
                combined_wb,
                report_type,
                files
            )
            if sheet_name:
                successful_sheets += 1

        if successful_sheets == 0:
            logger.error("No sheets could be created for the combined report")
            return None

        # Generate filename with timestamp
        timestamp = get_current_timestamp(TIMESTAMP_FORMAT)
        combined_filename = f"Combined_All_Reports_{timestamp}.xlsx"

        # Determine save path with PyInstaller support
        # For frozen apps, save to external directory (next to .exe)
        # For development, save to base path
        if is_frozen():
            # Save to external directory where reports are downloaded
            reports_dir_path = get_executable_dir() / REPORTS_DIR
        else:
            # Development mode: use base path
            reports_dir_path = get_base_path() / REPORTS_DIR

        # Ensure directory exists
        reports_dir_path.mkdir(parents=True, exist_ok=True)
        combined_file_path = reports_dir_path / combined_filename

        # Save the combined workbook
        combined_wb.save(combined_file_path)
        logger.debug(f"Saved combined report to: {combined_file_path}")
        logger.info(f"Successfully created combined report: {combined_filename}")
        logger.info(
            f"Combined {successful_sheets} report type sheets into single file"
        )

        return str(combined_file_path)

    except Exception as e:
        logger.error(f"Error combining reports: {e}")
        return None


def get_school_summary(
    directory: Union[str, Path]
) -> Optional[Dict[str, Any]]:
    """Get summary data for school overview from Student Usage reports.

    Args:
        directory: Path to the reports directory

    Returns:
        Dict with summary data or None if no data found
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find Student Usage files
        pattern = "*_Student Usage_*.xlsx"
        files = list(reports_directory.glob(pattern))

        if not files:
            logger.info("No Student Usage reports found")
            return None

        logger.info(f"Found {len(files)} Student Usage reports")

        # Process all files
        all_data = []
        for file_path in sorted(files):
            try:
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                df = pd.read_excel(file_path, engine="openpyxl")

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Collect data rows
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))
                    all_data.append(cleaned_row)

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        if not all_data:
            logger.info("No data found in Student Usage reports")
            return None

        # Calculate totals
        total_students = len(all_data)
        total_teachers = len(files)
        total_listens = 0
        total_reads = 0
        total_quizzes = 0

        for row_data in all_data:
            if (len(row_data) > STUDENT_USAGE_COL_LISTEN and
                    pd.notna(row_data[STUDENT_USAGE_COL_LISTEN])):
                try:
                    total_listens += float(row_data[STUDENT_USAGE_COL_LISTEN])
                except (ValueError, TypeError):
                    pass

            if (len(row_data) > STUDENT_USAGE_COL_READ and
                    pd.notna(row_data[STUDENT_USAGE_COL_READ])):
                try:
                    total_reads += float(row_data[STUDENT_USAGE_COL_READ])
                except (ValueError, TypeError):
                    pass

            if (len(row_data) > STUDENT_USAGE_COL_QUIZ and
                    pd.notna(row_data[STUDENT_USAGE_COL_QUIZ])):
                try:
                    total_quizzes += float(row_data[STUDENT_USAGE_COL_QUIZ])
                except (ValueError, TypeError):
                    pass

        total_activities = total_listens + total_reads + total_quizzes

        summary_data = {
            "all_teachers": total_teachers,
            "all_students": total_students,
            "total_listen": int(total_listens),
            "total_read": int(total_reads),
            "total_quizzes": int(total_quizzes),
            "total_activities": int(total_activities),
        }

        logger.debug(f"Calculated summary: {summary_data}")
        return summary_data

    except Exception as e:
        logger.error(f"Error getting school summary: {e}")
        return None


def get_classroom_summaries(
    directory: Union[str, Path]
) -> Optional[List[Dict[str, Any]]]:
    """Get per-classroom summary data from all Student Usage reports.

    Args:
        directory: Path to the reports directory

    Returns:
        List of classroom summary dictionaries or None if no data found
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find Student Usage files
        pattern = "*_Student Usage_*.xlsx"
        files = list(reports_directory.glob(pattern))

        if not files:
            logger.info("No Student Usage reports found")
            return None

        logger.info(
            f"Found {len(files)} Student Usage reports for classroom summaries"
        )

        # Dictionary to accumulate data per classroom
        classroom_data = {}

        for file_path in sorted(files):
            try:
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                # Extract classroom name from filename
                filename = file_path.name
                if "_Student Usage_" in filename:
                    classroom_name = filename.split("_Student Usage_")[0]
                else:
                    logger.warning(
                        f"Could not extract classroom name from {filename}, "
                        f"using 'Unknown'"
                    )
                    classroom_name = "Unknown"

                df = pd.read_excel(file_path, engine="openpyxl")

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Process each row in the file
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))

                    # Initialize classroom data if not exists
                    if classroom_name not in classroom_data:
                        classroom_data[classroom_name] = {
                            "name": classroom_name,
                            "students": 0,
                            "students_used": 0,
                            "listen": 0,
                            "read": 0,
                            "quiz": 0,
                            "interactivity": 0,
                            "practice_recording": 0,
                            "usage": 0
                        }

                    # Increment student count
                    classroom_data[classroom_name]["students"] += 1

                    # Check if student used the tool
                    student_used = False

                    # Sum listens
                    if (len(cleaned_row) > STUDENT_USAGE_COL_LISTEN and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_LISTEN])):
                        try:
                            listen_val = float(
                                cleaned_row[STUDENT_USAGE_COL_LISTEN]
                            )
                            classroom_data[classroom_name]["listen"] += listen_val
                            if listen_val > 0:
                                student_used = True
                        except (ValueError, TypeError):
                            pass

                    # Sum reads
                    if (len(cleaned_row) > STUDENT_USAGE_COL_READ and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_READ])):
                        try:
                            read_val = float(cleaned_row[STUDENT_USAGE_COL_READ])
                            classroom_data[classroom_name]["read"] += read_val
                            if read_val > 0:
                                student_used = True
                        except (ValueError, TypeError):
                            pass

                    # Sum quizzes
                    if (len(cleaned_row) > STUDENT_USAGE_COL_QUIZ and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_QUIZ])):
                        try:
                            quiz_val = float(cleaned_row[STUDENT_USAGE_COL_QUIZ])
                            classroom_data[classroom_name]["quiz"] += quiz_val
                            if quiz_val > 0:
                                student_used = True
                        except (ValueError, TypeError):
                            pass

                    # Count student as used if they have activity
                    if student_used:
                        classroom_data[classroom_name]["students_used"] += 1

                    # Sum interactivity
                    if (len(cleaned_row) > STUDENT_USAGE_COL_INTERACTIVITY and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_INTERACTIVITY])):
                        try:
                            classroom_data[classroom_name]["interactivity"] += float(
                                cleaned_row[STUDENT_USAGE_COL_INTERACTIVITY]
                            )
                        except (ValueError, TypeError):
                            pass

                    # Sum practice recording
                    if (len(cleaned_row) > STUDENT_USAGE_COL_PRACTICE_RECORDING and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_PRACTICE_RECORDING])):
                        try:
                            classroom_data[classroom_name]["practice_recording"] += float(
                                cleaned_row[STUDENT_USAGE_COL_PRACTICE_RECORDING]
                            )
                        except (ValueError, TypeError):
                            pass

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        if not classroom_data:
            logger.info("No classroom data found in Student Usage reports")
            return None

        # Calculate usage percentage for each classroom
        for classroom in classroom_data.values():
            total_students = classroom["students"]
            students_used = classroom["students_used"]
            if total_students > 0:
                classroom["usage"] = round(
                    (students_used / total_students) * 100, 1
                )
            else:
                classroom["usage"] = 0

        # Convert to list and sort by classroom name
        classroom_summaries = list(classroom_data.values())
        classroom_summaries.sort(key=lambda x: x["name"])

        logger.debug(
            f"Calculated summaries for {len(classroom_summaries)} classrooms"
        )
        return classroom_summaries

    except Exception as e:
        logger.error(f"Error getting classroom summaries: {e}")
        return None


def get_reading_skills_data(
    directory: Union[str, Path]
) -> Optional[List[Dict[str, Any]]]:
    """Get reading skills data from Skill reports for each classroom.

    Args:
        directory: Path to the reports directory

    Returns:
        List of classroom skill dictionaries or None if no data found
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find Skill files
        pattern = "*_Skill_*.xlsx"
        files = list(reports_directory.glob(pattern))

        if not files:
            logger.info("No Skill reports found")
            return None

        logger.info(f"Found {len(files)} Skill reports")

        # Dictionary to accumulate data per classroom
        classroom_data = {}

        for file_path in sorted(files):
            try:
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                # Extract classroom name from filename
                filename = file_path.name
                if "_Skill_" in filename:
                    classroom_name = filename.split("_Skill_")[0]
                else:
                    logger.warning(
                        f"Could not extract classroom name from {filename}, "
                        f"skipping"
                    )
                    continue

                df = pd.read_excel(file_path, engine="openpyxl")

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Initialize classroom data if not exists
                if classroom_name not in classroom_data:
                    classroom_data[classroom_name] = {
                        "classroom": classroom_name,
                        "skills": []
                    }

                # Process each row in the file
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))

                    # Extract skill data from columns
                    if len(cleaned_row) >= 4:
                        skill_data = {
                            "name": str(cleaned_row[SKILL_COL_SKILL_NAME]).strip(),
                            "correct": (
                                int(cleaned_row[SKILL_COL_CORRECT])
                                if pd.notna(cleaned_row[SKILL_COL_CORRECT]) and
                                str(cleaned_row[SKILL_COL_CORRECT]).isdigit()
                                else 0
                            ),
                            "total": (
                                int(cleaned_row[SKILL_COL_TOTAL])
                                if pd.notna(cleaned_row[SKILL_COL_TOTAL]) and
                                str(cleaned_row[SKILL_COL_TOTAL]).isdigit()
                                else 0
                            ),
                            "accuracy": (
                                float(cleaned_row[SKILL_COL_ACCURACY])
                                if pd.notna(cleaned_row[SKILL_COL_ACCURACY])
                                else 0.0
                            )
                        }

                        # Calculate accuracy if not provided
                        if (skill_data["total"] > 0 and
                                skill_data["accuracy"] == 0.0):
                            skill_data["accuracy"] = round(
                                (skill_data["correct"] / skill_data["total"]) * 100,
                                1
                            )

                        classroom_data[classroom_name]["skills"].append(skill_data)

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        if not classroom_data:
            logger.info("No classroom data found in Skill reports")
            return None

        # Convert to list and sort by classroom name
        reading_skills_data = list(classroom_data.values())
        reading_skills_data.sort(key=lambda x: x["classroom"])

        logger.debug(
            f"Calculated skills data for {len(reading_skills_data)} classrooms"
        )
        return reading_skills_data

    except Exception as e:
        logger.error(f"Error getting reading skills data: {e}")
        return None


def get_skills_summary(
    directory: Union[str, Path]
) -> Optional[Dict[str, Any]]:
    """Get aggregated skills summary across all classrooms.

    Args:
        directory: Path to the reports directory

    Returns:
        Dictionary with totals and averages or None if no data found
    """
    try:
        # Get per-classroom skills data
        reading_skills_data = get_reading_skills_data(directory)

        if not reading_skills_data:
            logger.info("No reading skills data found for summary")
            return None

        # Initialize accumulators
        total_classrooms = len(reading_skills_data)
        total_correct = 0
        total_questions = 0

        # Dictionary to accumulate per-skill data for averaging
        skill_accumulator: Dict[str, Dict[str, Any]] = {}

        # Process each classroom
        for classroom_data in reading_skills_data:
            skills = classroom_data.get("skills", [])

            for skill in skills:
                correct = skill.get("correct", 0)
                total = skill.get("total", 0)
                accuracy = skill.get("accuracy", 0.0)
                skill_name = skill.get("name", "Unknown")

                total_correct += correct
                total_questions += total

                # Accumulate for per-skill averaging
                if skill_name not in skill_accumulator:
                    skill_accumulator[skill_name] = {
                        "total_accuracy": 0.0,
                        "count": 0,
                        "total_correct": 0,
                        "total_questions": 0
                    }

                skill_accumulator[skill_name]["total_accuracy"] += accuracy
                skill_accumulator[skill_name]["count"] += 1
                skill_accumulator[skill_name]["total_correct"] += correct
                skill_accumulator[skill_name]["total_questions"] += total

        # Calculate overall accuracy
        overall_accuracy = 0.0
        if total_questions > 0:
            overall_accuracy = round((total_correct / total_questions) * 100, 1)

        # Calculate per-skill averages and sort by accuracy
        skill_averages = []
        for skill_name, data in skill_accumulator.items():
            if data["count"] > 0:
                avg_accuracy = round(data["total_accuracy"] / data["count"], 1)
                skill_averages.append({
                    "name": skill_name,
                    "accuracy": avg_accuracy,
                    "correct": data["total_correct"],
                    "total": data["total_questions"]
                })

        # Sort by accuracy descending
        skill_averages.sort(key=lambda x: x["accuracy"], reverse=True)

        # Count skills by accuracy category for pie chart
        high_accuracy_count = sum(
            1 for s in skill_averages if s["accuracy"] >= 80
        )
        medium_accuracy_count = sum(
            1 for s in skill_averages if 60 <= s["accuracy"] < 80
        )
        low_accuracy_count = sum(
            1 for s in skill_averages if s["accuracy"] < 60
        )

        summary = {
            "total_classrooms": total_classrooms,
            "total_correct": total_correct,
            "total_questions": total_questions,
            "overall_accuracy": overall_accuracy,
            "skill_averages": skill_averages,
            "top_skills": skill_averages[:5],
            "accuracy_distribution": {
                "high": high_accuracy_count,
                "medium": medium_accuracy_count,
                "low": low_accuracy_count
            }
        }

        logger.debug(
            f"Calculated skills summary: {total_classrooms} classrooms, "
            f"{overall_accuracy}% accuracy"
        )
        return summary

    except Exception as e:
        logger.error(f"Error getting skills summary: {e}")
        return None


def get_top_readers_per_classroom(
    directory: Union[str, Path]
) -> Optional[List[Dict[str, Any]]]:
    """Get top 3 readers per classroom from Student Usage reports.

    Args:
        directory: Path to the reports directory

    Returns:
        List of classroom dictionaries with top readers or None if no data found
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find Student Usage files
        pattern = "*_Student Usage_*.xlsx"
        files = list(reports_directory.glob(pattern))

        if not files:
            logger.info("No Student Usage reports found")
            return None

        logger.info(f"Found {len(files)} Student Usage reports for top readers")

        # Dictionary to accumulate students per classroom
        classroom_students = {}

        for file_path in sorted(files):
            try:
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                # Extract classroom name from filename
                filename = file_path.name
                if "_Student Usage_" in filename:
                    classroom_name = filename.split("_Student Usage_")[0]
                else:
                    logger.warning(
                        f"Could not extract classroom name from {filename}, "
                        f"using 'Unknown'"
                    )
                    classroom_name = "Unknown"

                df = pd.read_excel(file_path, engine="openpyxl")

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Process each row in the file
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))

                    # Extract student name
                    student_name = (
                        str(cleaned_row[STUDENT_USAGE_COL_STUDENT_NAME]).strip()
                        if len(cleaned_row) > STUDENT_USAGE_COL_STUDENT_NAME
                        else ""
                    )

                    if not student_name:
                        continue

                    # Extract listen and read values
                    listen = 0
                    read = 0

                    if (len(cleaned_row) > STUDENT_USAGE_COL_LISTEN and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_LISTEN])):
                        try:
                            listen = float(cleaned_row[STUDENT_USAGE_COL_LISTEN])
                        except (ValueError, TypeError):
                            pass

                    if (len(cleaned_row) > STUDENT_USAGE_COL_READ and
                            pd.notna(cleaned_row[STUDENT_USAGE_COL_READ])):
                        try:
                            read = float(cleaned_row[STUDENT_USAGE_COL_READ])
                        except (ValueError, TypeError):
                            pass

                    score = listen + read

                    # Initialize classroom if not exists
                    if classroom_name not in classroom_students:
                        classroom_students[classroom_name] = []

                    # Add student to classroom
                    classroom_students[classroom_name].append({
                        "name": student_name,
                        "score": score
                    })

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        if not classroom_students:
            logger.info("No student data found in Student Usage reports")
            return None

        # Process each classroom to get top 3 readers
        top_readers_data = []

        for classroom_name, students in classroom_students.items():
            # Filter out students with zero scores
            valid_students = [s for s in students if s["score"] > 0]

            # Skip classroom if all students have zero scores
            if not valid_students:
                continue

            # Sort by score descending and take top 3
            valid_students.sort(key=lambda x: x["score"], reverse=True)
            top_students = valid_students[:3]

            top_readers_data.append({
                "name": classroom_name,
                "students": top_students
            })

        # Sort classrooms by name
        top_readers_data.sort(key=lambda x: x["name"])

        logger.debug(
            f"Calculated top readers for {len(top_readers_data)} classrooms"
        )
        return top_readers_data

    except Exception as e:
        logger.error(f"Error getting top readers per classroom: {e}")
        return None


def get_level_up_progress_data(
    directory: Union[str, Path]
) -> Optional[List[Dict[str, Any]]]:
    """Get level up progress data from Level Up Progress reports.

    Args:
        directory: Path to the reports directory

    Returns:
        List of classroom level up dictionaries or None if no data found
    """
    try:
        reports_directory = _resolve_reports_directory(directory)
        if not reports_directory:
            return None

        # Find Level Up Progress files
        pattern = "*_Level Up Progress_*.xlsx"
        files = list(reports_directory.glob(pattern))

        if not files:
            logger.info("No Level Up Progress reports found")
            return None

        logger.info(f"Found {len(files)} Level Up Progress reports")

        # Dictionary to accumulate data per classroom
        classroom_data = {}

        for file_path in sorted(files):
            try:
                if not file_path.exists() or file_path.stat().st_size == 0:
                    logger.warning(
                        f"File {file_path} is empty or doesn't exist, skipping"
                    )
                    continue

                # Extract classroom name from filename
                filename = file_path.name
                if "_Level Up Progress_" in filename:
                    classroom_name = filename.split("_Level Up Progress_")[0]
                else:
                    logger.warning(
                        f"Could not extract classroom name from {filename}, "
                        f"skipping"
                    )
                    continue

                df = pd.read_excel(file_path, engine="openpyxl")

                if df.empty:
                    logger.warning(
                        f"Excel file {file_path} contains no data, skipping"
                    )
                    continue

                # Initialize classroom data if not exists
                if classroom_name not in classroom_data:
                    classroom_data[classroom_name] = {
                        "classroom": classroom_name,
                        "students": []
                    }

                # Process each row in the file
                for _, row in df.iterrows():
                    cleaned_row = clean_row_data(list(row))

                    # Extract level up data from columns
                    if len(cleaned_row) >= 8:
                        # Parse progress value (remove % and convert to float)
                        progress_str = str(
                            cleaned_row[LEVEL_UP_COL_PROGRESS]
                        ).strip()
                        if "%" in progress_str:
                            progress_str = progress_str.replace("%", "")
                        try:
                            progress_value = float(progress_str)
                        except (ValueError, TypeError):
                            progress_value = 0.0

                        student_data = {
                            "student": str(
                                cleaned_row[LEVEL_UP_COL_STUDENT]
                            ).strip(),
                            "level": str(cleaned_row[LEVEL_UP_COL_LEVEL]).strip(),
                            "progress": progress_value
                        }

                        classroom_data[classroom_name]["students"].append(
                            student_data
                        )

            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                continue

        if not classroom_data:
            logger.info("No classroom data found in Level Up Progress reports")
            return None

        # Convert to list and sort by classroom name
        level_up_progress_data = list(classroom_data.values())
        # Sort students alphabetically by name within each classroom
        for classroom in level_up_progress_data:
            classroom["students"].sort(key=lambda x: x["student"])

        # Sort by classroom name
        level_up_progress_data.sort(key=lambda x: x["classroom"])

        logger.debug(
            f"Calculated level up progress data for "
            f"{len(level_up_progress_data)} classrooms"
        )
        return level_up_progress_data

    except Exception as e:
        logger.error(f"Error getting level up progress data: {e}")
        return None


def get_classroom_comparison_data(
    directory: Union[str, Path]
) -> Optional[Dict[str, Any]]:
    """Get classroom comparison data for bar charts.

    Args:
        directory: Path to the reports directory

    Returns:
        Dict with labels and data for listen, read, quiz charts or None
    """
    summaries = get_classroom_summaries(directory)

    if not summaries:
        return None

    labels = [c["name"] for c in summaries]
    listen_data = [c["listen"] for c in summaries]
    read_data = [c["read"] for c in summaries]
    quiz_data = [c["quiz"] for c in summaries]

    return {
        "labels": labels,
        "listen": listen_data,
        "read": read_data,
        "quiz": quiz_data
    }
