"""Excel file manipulation utilities for the reports scraper application.

This module provides functions for Excel styling, formatting, data bars,
and worksheet manipulation to reduce code duplication.
"""

import logging
from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import DataBarRule

from constants import (
    EXCEL_HEADER_COLOR,
    EXCEL_HEADER_FONT_COLOR,
    EXCEL_DATA_BAR_COLOR,
    EXCEL_MAX_SHEET_NAME_LENGTH,
    EXCEL_INVALID_SHEET_CHARS,
    EXCEL_MIN_COLUMN_WIDTH,
    EXCEL_MAX_COLUMN_WIDTH
)

logger = logging.getLogger(__name__)


def get_header_fill() -> PatternFill:
    """Get the standard header fill pattern.

    Returns:
        PatternFill for header cells (light blue background)
    """
    return PatternFill(
        start_color=EXCEL_HEADER_COLOR,
        end_color=EXCEL_HEADER_COLOR,
        fill_type="solid"
    )


def get_header_font() -> Font:
    """Get the standard header font.

    Returns:
        Font for header cells (bold black text)
    """
    return Font(bold=True, color=EXCEL_HEADER_FONT_COLOR)


def get_header_alignment() -> Alignment:
    """Get the standard header alignment.

    Returns:
        Alignment for header cells (centered)
    """
    return Alignment(horizontal="center", vertical="center")


def apply_header_style(
    worksheet: Worksheet,
    row_num: int,
    start_col: Optional[int] = None,
    end_col: Optional[int] = None
) -> None:
    """Apply header styling to a row.

    Args:
        worksheet: Worksheet to modify
        row_num: Row number to style (1-based)
        start_col: Starting column index (None for all columns)
        end_col: Ending column index (None for all columns)
    """
    try:
        header_fill = get_header_fill()
        header_font = get_header_font()
        header_alignment = get_header_alignment()

        row = worksheet[row_num]

        # Determine which cells to style
        if start_col is not None and end_col is not None:
            cells = row[start_col:end_col + 1]
        else:
            cells = row

        for cell in cells:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        logger.debug(f"Applied header styling to row {row_num}")

    except Exception as e:
        logger.warning(f"Error applying header style to row {row_num}: {e}")


def auto_adjust_column_widths(
    worksheet: Worksheet,
    min_width: int = EXCEL_MIN_COLUMN_WIDTH,
    max_width: int = EXCEL_MAX_COLUMN_WIDTH
) -> None:
    """Auto-adjust column widths based on content.

    Args:
        worksheet: Worksheet to modify
        min_width: Minimum column width
        max_width: Maximum column width
    """
    try:
        for column_cells in worksheet.columns:
            # Calculate max length in column
            length = max(
                len(str(cell.value or ''))
                for cell in column_cells
            )

            # Set width with padding and limits
            adjusted_width = min(length + 2, max_width)
            adjusted_width = max(adjusted_width, min_width)

            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.debug("Auto-adjusted column widths")

    except Exception as e:
        logger.warning(f"Error auto-adjusting column widths: {e}")


def add_data_bars(
    worksheet: Worksheet,
    column_letter: str,
    start_row: int,
    end_row: int,
    min_value: int = 0,
    max_value: int = 100,
    color: str = EXCEL_DATA_BAR_COLOR
) -> None:
    """Add data bars to a column range.

    Args:
        worksheet: Worksheet to modify
        column_letter: Column letter (e.g., 'D')
        start_row: Starting row number (1-based)
        end_row: Ending row number (1-based)
        min_value: Minimum value for data bar scale
        max_value: Maximum value for data bar scale
        color: Hex color code for data bars
    """
    try:
        if start_row >= end_row:
            logger.warning(
                f"Invalid row range for data bars: {start_row} to {end_row}"
            )
            return

        data_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

        data_bar_rule = DataBarRule(
            start_type='num',
            start_value=min_value,
            end_type='num',
            end_value=max_value,
            color=color,
            showValue=True,
            minLength=0,
            maxLength=100
        )

        worksheet.conditional_formatting.add(data_range, data_bar_rule)
        logger.debug(f"Added data bars to range {data_range}")

    except Exception as e:
        logger.warning(f"Error adding data bars: {e}")


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a sheet name to be valid for Excel.

    Excel has restrictions: max 31 chars, no special chars like : \\ / ? * [ ]

    Args:
        name: Proposed sheet name

    Returns:
        Sanitized sheet name
    """
    # Truncate to max length
    sanitized = name[:EXCEL_MAX_SHEET_NAME_LENGTH]

    # Remove invalid characters
    for char in EXCEL_INVALID_SHEET_CHARS:
        sanitized = sanitized.replace(char, '_')

    return sanitized


def make_unique_sheet_name(
    proposed_name: str,
    existing_names: List[str]
) -> str:
    """Create a unique sheet name.

    If the proposed name already exists, appends a number.

    Args:
        proposed_name: Desired sheet name
        existing_names: List of existing sheet names

    Returns:
        Unique sheet name
    """
    sanitized = sanitize_sheet_name(proposed_name)

    if sanitized not in existing_names:
        return sanitized

    # Append counter until unique
    counter = 1
    while True:
        # Leave room for counter (e.g., "_1", "_99")
        base_name = sanitized[:EXCEL_MAX_SHEET_NAME_LENGTH - 3]
        unique_name = f"{base_name}_{counter}"

        if unique_name not in existing_names:
            return unique_name

        counter += 1

        # Safety check to prevent infinite loop
        if counter > 999:
            logger.error(
                f"Could not create unique name for '{proposed_name}'"
            )
            return f"{base_name}_X"


def create_summary_row(
    label: str,
    value: any,
    num_columns: int
) -> List[any]:
    """Create a summary row with label in column A, value in column B.

    Args:
        label: Summary row label
        value: Summary value
        num_columns: Total number of columns in the sheet

    Returns:
        List representing the row with empty cells for other columns
    """
    row = [''] * num_columns
    row[0] = label
    row[1] = value
    return row


def style_summary_row(
    worksheet: Worksheet,
    row_num: int
) -> None:
    """Apply styling to a summary row (columns A and B only).

    Args:
        worksheet: Worksheet to modify
        row_num: Row number to style (1-based)
    """
    try:
        header_fill = get_header_fill()
        header_font = get_header_font()

        # Style column A (label)
        cell_a = worksheet[f'A{row_num}']
        cell_a.fill = header_fill
        cell_a.font = header_font
        cell_a.alignment = Alignment(horizontal="left", vertical="center")

        # Style column B (value)
        cell_b = worksheet[f'B{row_num}']
        cell_b.fill = header_fill
        cell_b.font = header_font
        cell_b.alignment = Alignment(horizontal="right", vertical="center")

        logger.debug(f"Styled summary row {row_num}")

    except Exception as e:
        logger.warning(f"Error styling summary row {row_num}: {e}")


def create_separator_row(
    username: str,
    num_columns: int
) -> List[str]:
    """Create a separator row with username.

    Args:
        username: Username to display
        num_columns: Total number of columns in the sheet

    Returns:
        List representing the separator row
    """
    row = [''] * num_columns
    row[0] = f"User: {username}"
    return row


def clean_row_data(row_data: List[any]) -> List[any]:
    """Clean row data by replacing None/NaN values with empty strings.

    Args:
        row_data: List of cell values

    Returns:
        Cleaned list with no None/NaN values
    """
    import pandas as pd
    return [cell if pd.notna(cell) else '' for cell in row_data]
